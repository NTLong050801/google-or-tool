"""Export TKB theo lớp: 1 file Excel, mỗi lớp 1 sheet (lưới Thứ × Tiết)."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schemas import Assignment, Classroom, ScheduledSession

_DAY_VI = {2: "Thứ 2", 3: "Thứ 3", 4: "Thứ 4", 5: "Thứ 5", 6: "Thứ 6", 7: "Thứ 7", 8: "Chủ nhật"}

_PROGRAM_LABEL = {
    "trung_cap": "Trung cấp",
    "cao_dang": "Cao đẳng",
    "lien_thong": "Liên thông",
}

_REASON_LABELS: dict = {
    "thi":      "Tuần thi",
    "du_phong": "Dự phòng",
    "quan_su":  "Quân sự",
    "nghi_le":  "Nghỉ lễ",
    "thi_lai":  "Thi lại",
    "nghi":     "Nghỉ",
    "thuc_te":  "Thực tế/Thực tập",
    "khac":     "Khác",
}

_THIN = Side(style="thin", color="888888")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_MORNING_FILL = PatternFill("solid", fgColor="FFF2CC")
_AFTERNOON_FILL = PatternFill("solid", fgColor="E2EFDA")
_CONFLICT_FILL = PatternFill("solid", fgColor="F8CBAD")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Màu cho bảng tóm tắt tuần
_WEEK_HEADER_FILL  = PatternFill("solid", fgColor="2F4F8F")   # xanh đậm — header
_WEEK_STUDY_FILL   = PatternFill("solid", fgColor="E2EFDA")   # xanh nhạt — tuần học
_WEEK_HOLIDAY_FILL = PatternFill("solid", fgColor="FCE4D6")   # cam nhạt — nghỉ lễ
_WEEK_EXAM_FILL    = PatternFill("solid", fgColor="FFF2CC")   # vàng — thi
_WEEK_RESERVE_FILL = PatternFill("solid", fgColor="EDEDED")   # xám — dự phòng
_WEEK_MILITARY_FILL= PatternFill("solid", fgColor="D6E4F0")   # xanh dương nhạt — quân sự
_WEEK_OTHER_FILL   = PatternFill("solid", fgColor="F2F2F2")   # xám nhạt — khác

_REASON_TO_FILL = {
    "nghi_le":  _WEEK_HOLIDAY_FILL,
    "nghi":     _WEEK_HOLIDAY_FILL,
    "thi":      _WEEK_EXAM_FILL,
    "thi_lai":  _WEEK_EXAM_FILL,
    "du_phong": _WEEK_RESERVE_FILL,
    "quan_su":  _WEEK_MILITARY_FILL,
    "thuc_te":  _WEEK_OTHER_FILL,
    "khac":     _WEEK_OTHER_FILL,
}


def _write_week_summary(
    ws,
    start_row: int,
    cls_id: str,
    week_dates: Dict[int, Tuple[str, str]],
    holiday_reasons: Dict[int, str],
    class_excluded_weeks: Dict[str, Dict[int, str]],
    class_week_start: Optional[int],
    class_week_end: Optional[int],
    n_cols: int,
) -> None:
    """Vẽ bảng tóm tắt tuần phía dưới lưới TKB.

    Mỗi hàng = 1 tuần. Cột: STT tuần | Ngày bắt đầu | Ngày kết thúc | Trạng thái | Ghi chú.
    Tuần học  → xanh nhạt. Tuần nghỉ/thi/dự phòng → màu riêng.
    Tuần ngoài phạm vi lớp (trước week_start) → xám, label "Chưa bắt đầu".
    """
    class_excl = class_excluded_weeks.get(cls_id, {})

    # Header bảng tóm tắt
    header_row = start_row
    header_font = Font(bold=True, color="FFFFFF")
    header_cols = ["Tuần", "Từ ngày", "Đến ngày", "Trạng thái", "Ghi chú"]

    title_cell = ws.cell(row=header_row - 1, column=1, value="LỊCH TUẦN HỌC KỲ")
    title_cell.font = Font(bold=True, size=11)
    title_cell.alignment = _LEFT
    if n_cols > 1:
        ws.merge_cells(
            start_row=header_row - 1, start_column=1,
            end_row=header_row - 1, end_column=min(n_cols, 5),
        )

    for ci, col_name in enumerate(header_cols):
        c = ws.cell(row=header_row, column=1 + ci, value=col_name)
        c.font = header_font
        c.fill = _WEEK_HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    # Một hàng mỗi tuần, theo thứ tự week_order
    for ri, week_order in enumerate(sorted(week_dates.keys()), start=1):
        from_date, to_date = week_dates[week_order]
        row = header_row + ri

        # Xác định trạng thái và màu
        if week_order in holiday_reasons:
            reason_key = holiday_reasons[week_order]
            status = _REASON_LABELS.get(reason_key, reason_key)
            note = ""
            fill = _REASON_TO_FILL.get(reason_key, _WEEK_OTHER_FILL)
        elif week_order in class_excl:
            reason_key = class_excl[week_order]
            status = _REASON_LABELS.get(reason_key, reason_key)
            note = "(riêng lớp)"
            fill = _REASON_TO_FILL.get(reason_key, _WEEK_OTHER_FILL)
        elif class_week_start is not None and week_order < class_week_start:
            status = "Chưa bắt đầu"
            note = ""
            fill = _WEEK_RESERVE_FILL
        elif class_week_end is not None and week_order > class_week_end:
            status = "Kết thúc"
            note = ""
            fill = _WEEK_RESERVE_FILL
        else:
            status = "Học"
            note = ""
            fill = _WEEK_STUDY_FILL

        row_data = [
            f"Tuần {week_order}",
            _format_date(from_date),
            _format_date(to_date),
            status,
            note,
        ]
        for ci, val in enumerate(row_data):
            c = ws.cell(row=row, column=1 + ci, value=val)
            c.fill = fill
            c.border = _BORDER
            c.alignment = _CENTER

    # Điều chỉnh độ rộng cột bảng tóm tắt
    col_widths = [10, 13, 13, 16, 14]
    for ci, w in enumerate(col_widths):
        col_letter = get_column_letter(1 + ci)
        current = ws.column_dimensions[col_letter].width or 0
        ws.column_dimensions[col_letter].width = max(current, w)


def _safe_sheet_name(raw: str, used: set) -> str:
    """Excel sheet name: max 31 chars, không chứa : \\ / ? * [ ]."""
    s = raw or "lop"
    for ch in r":\/?*[]":
        s = s.replace(ch, "_")
    s = s[:31].strip() or "lop"
    base = s
    n = 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = (base[: 31 - len(suffix)]) + suffix
    used.add(s)
    return s


def _format_date(s: str) -> str:
    """YYYY-MM-DD → DD/MM/YYYY (rút gọn cho hiển thị)."""
    if not s or len(s) < 10:
        return s
    try:
        y, m, d = s[:10].split("-")
        return f"{d}/{m}/{y}"
    except ValueError:
        return s


def _format_week_ranges(weeks: list[int]) -> str:
    """Chuyển danh sách tuần thành chuỗi range: [24,26,27,28,31] → '24, 26-28, 31'."""
    if not weeks:
        return ""
    ranges = []
    start = end = weeks[0]
    for w in weeks[1:]:
        if w == end + 1:
            end = w
        else:
            ranges.append(str(start) if start == end else f"{start}-{end}")
            start = end = w
    ranges.append(str(start) if start == end else f"{start}-{end}")
    return ", ".join(ranges)


def _build_cell_text(
    s: ScheduledSession,
    a: Optional[Assignment],
    label: Dict[str, str],
    classrooms_by_id: Dict[int, Classroom],
    week_dates: Dict[int, Tuple[str, str]],
    holiday_reasons: Dict[int, str],
    class_excluded_weeks: Dict[str, Dict[int, str]],
) -> str:
    sub_code = label.get("subject_code", "")
    sub_name = label.get("subject_name", "")
    total_hours = label.get("total_hours", "")
    total_sessions = label.get("total_sessions", "")
    teacher_name = label.get("teacher_name", s.teacher_id)
    teacher_type = label.get("teacher_type", "").strip()
    cluster = int(a.lessons_cluster) if a else (int(s.period_end) - int(s.period_start) + 1)

    room = classrooms_by_id.get(int(s.classroom_id))
    room_label = (room.name if room else None) or str(s.classroom_id)

    w_start = int(s.week_start)
    w_end = int(s.week_end)
    fd_start, _ = week_dates.get(w_start, ("", ""))
    _, td_end = week_dates.get(w_end, ("", ""))

    # Dùng teaching_weeks/skipped_weeks từ solver nếu có (đã tính sẵn, nhất quán với solver)
    # Fallback: tự tính từ holiday_reasons + class_excluded_weeks (backward compat)
    if s.teaching_weeks:
        teaching_weeks = s.teaching_weeks
        all_skipped: Dict[int, str] = dict(s.skipped_weeks)
    else:
        cls_id = label.get("class_id", "")
        class_excl = class_excluded_weeks.get(cls_id, {})
        excl_in_range = {w: r for w, r in class_excl.items() if w_start <= w <= w_end}
        holidays_in_range = {
            w: holiday_reasons.get(w, "Nghỉ")
            for w in range(w_start, w_end + 1)
            if w in holiday_reasons
        }
        all_skipped = {**holidays_in_range, **excl_in_range}
        teaching_weeks = sorted(w for w in range(w_start, w_end + 1) if w not in all_skipped)
    teaching_range = _format_week_ranges(teaching_weeks)

    # Dòng tuần: chỉ hiển thị tuần thực dạy kèm ngày tháng
    if teaching_weeks:
        fd_teach, _ = week_dates.get(teaching_weeks[0], ("", ""))
        _, td_teach = week_dates.get(teaching_weeks[-1], ("", ""))
        week_line = f"Tuần {teaching_range}"
        if fd_teach and td_teach:
            week_line += f" ({_format_date(fd_teach)} → {_format_date(td_teach)})"
        week_line += f"  [{len(teaching_weeks)} tuần thực dạy]"
    else:
        week_line = f"Tuần {w_start}-{w_end}  [0 tuần thực dạy]"

    # Dòng các tuần không dạy (gom chung, kèm lý do)
    skipped_text = ""
    if all_skipped:
        skipped_items = [
            f"W{w} ({_REASON_LABELS.get(r, r)})"
            for w, r in sorted(all_skipped.items())
        ]
        skipped_text = "Không dạy: " + " | ".join(skipped_items)

    teacher_line = f"GV: {teacher_name}"
    if teacher_type:
        teacher_line += f" [{teacher_type}]"

    parts = [
        f"{sub_code} – {sub_name}" if sub_code else sub_name,
        f"Tổng: {total_sessions} buổi ({total_hours}h) · {cluster} tiết/buổi",
        teacher_line,
        f"Phòng: {room_label}",
        week_line,
    ]
    if skipped_text:
        parts.append(skipped_text)

    return "\n".join(p for p in parts if p)


def export_timetable_by_class(
    sessions: List[ScheduledSession],
    assignments: List[Assignment],
    assignment_labels: Dict[int, Dict[str, str]],
    classrooms: List[Classroom],
    *,
    days: List[int],
    periods_per_day: int,
    morning_periods: Optional[List[int]],
    xlsx_path: Path,
    week_dates: Optional[Dict[int, Tuple[str, str]]] = None,
    holiday_reasons: Optional[Dict[int, str]] = None,
    class_excluded_weeks: Optional[Dict[str, Dict[int, str]]] = None,
) -> None:
    """Xuất Excel: mỗi sheet = 1 lớp, lưới hàng=Tiết × cột=Thứ. Slot 5 tiết → merge."""
    a_by_id = {int(a.id): a for a in assignments}
    room_by_id = {int(c.id): c for c in classrooms}
    week_dates = week_dates or {}
    holiday_reasons = holiday_reasons or {}
    class_excluded_weeks = class_excluded_weeks or {}

    by_class: Dict[str, List[ScheduledSession]] = {}
    for s in sessions:
        lab = assignment_labels.get(int(s.assignment_id), {})
        cls = str(lab.get("class_id", "")).strip() or f"unknown_{s.assignment_id}"
        by_class.setdefault(cls, []).append(s)

    if not by_class:
        return

    last_morning = max(morning_periods) if morning_periods else periods_per_day // 2

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()

    for cls_name in sorted(by_class.keys()):
        cls_sessions = by_class[cls_name]
        sheet_name = _safe_sheet_name(cls_name, used_names)
        ws = wb.create_sheet(title=sheet_name)

        # Lấy program_level từ session đầu tiên
        first_label = assignment_labels.get(int(cls_sessions[0].assignment_id), {})
        program = first_label.get("program_level", "")
        program_label = _PROGRAM_LABEL.get(program, program or "")
        title = f"Thời khóa biểu lớp: {cls_name}"
        if program_label:
            title += f" ({program_label})"

        n_cols = 1 + len(days)
        ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = _LEFT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

        header_row = 3
        ws.cell(row=header_row, column=1, value="Tiết").font = Font(bold=True)
        ws.cell(row=header_row, column=1).fill = _HEADER_FILL
        ws.cell(row=header_row, column=1).alignment = _CENTER
        ws.cell(row=header_row, column=1).border = _BORDER
        for i, d in enumerate(days):
            c = ws.cell(row=header_row, column=2 + i, value=_DAY_VI.get(int(d), str(d)))
            c.font = Font(bold=True)
            c.fill = _HEADER_FILL
            c.alignment = _CENTER
            c.border = _BORDER

        for p in range(1, periods_per_day + 1):
            r = header_row + p
            cell = ws.cell(row=r, column=1, value=f"Tiết {p}")
            cell.alignment = _CENTER
            cell.border = _BORDER
            cell.fill = _MORNING_FILL if p <= last_morning else _AFTERNOON_FILL

        for p in range(1, periods_per_day + 1):
            r = header_row + p
            for i in range(len(days)):
                c = ws.cell(row=r, column=2 + i)
                c.border = _BORDER
                c.alignment = _CENTER
                c.fill = _MORNING_FILL if p <= last_morning else _AFTERNOON_FILL

        # Đặt nội dung từng session
        slot_content: Dict[Tuple[int, int], List[str]] = {}
        slot_span: Dict[Tuple[int, int], int] = {}
        for s in cls_sessions:
            a = a_by_id.get(int(s.assignment_id))
            cluster = int(a.lessons_cluster) if a else (int(s.period_end) - int(s.period_start) + 1)
            lab = assignment_labels.get(int(s.assignment_id), {})
            content = _build_cell_text(s, a, lab, room_by_id, week_dates, holiday_reasons, class_excluded_weeks)
            key = (int(s.day), int(s.period_start))
            slot_content.setdefault(key, []).append(content)
            slot_span[key] = cluster

        for (day, p_start), contents in slot_content.items():
            if day not in days:
                continue
            col = 2 + days.index(day)
            row = header_row + p_start
            cluster = slot_span[(day, p_start)]
            text = "\n──────\n".join(contents)
            cell = ws.cell(row=row, column=col, value=text)
            cell.alignment = _CENTER
            cell.border = _BORDER
            if len(contents) > 1:
                cell.fill = _CONFLICT_FILL

            if cluster > 1:
                ws.merge_cells(
                    start_row=row,
                    start_column=col,
                    end_row=row + cluster - 1,
                    end_column=col,
                )
                for r2 in range(row, row + cluster):
                    ws.cell(row=r2, column=col).border = _BORDER

        ws.column_dimensions["A"].width = 8
        for i in range(len(days)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 34
        for p in range(1, periods_per_day + 1):
            ws.row_dimensions[header_row + p].height = 30
        ws.row_dimensions[header_row].height = 18

        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

        # --- Bảng tóm tắt tuần học kỳ ---
        if week_dates:
            # week_start của lớp: lấy từ sessions (tuần thực dạy sớm nhất)
            cls_week_starts = [int(s.week_start) for s in cls_sessions]
            cls_week_start  = min(cls_week_starts) if cls_week_starts else None
            # week_end: KHÔNG lấy từ sessions (sessions chỉ có tuần học, không có tuần thi/nghỉ)
            # → dùng None để hiển thị toàn bộ tuần kỳ học với trạng thái đúng từ excluded_weeks
            cls_week_end = None

            # Khoảng cách 2 hàng trống sau lưới TKB
            summary_start_row = header_row + periods_per_day + 3
            _write_week_summary(
                ws,
                start_row=summary_start_row,
                cls_id=cls_name,
                week_dates=week_dates,
                holiday_reasons=holiday_reasons,
                class_excluded_weeks=class_excluded_weeks,
                class_week_start=cls_week_start,
                class_week_end=cls_week_end,
                n_cols=n_cols,
            )

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
