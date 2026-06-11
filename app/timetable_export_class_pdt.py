"""Export TKB theo mẫu phòng đào tạo: 1 file/khoa, mỗi lớp 1 sheet.

Layout (mỗi sheet):
  Row 1-2:   Header trường + CHXHCN + KHOA + tiêu đề (compact 3 dòng)
  Row 4-8:   Calendar 5 dòng (Tháng | Tuần | Từ ngày | Đến ngày | Trạng thái)
  Row 10:    Banner thi (auto từ class_excluded_weeks)
  Row 11-13: Header grid (TÊN LỚP | BUỔI | THỨ 2..CN × Tiết 1..5)
  Row 14+:   Data: 1 lớp × N sub-row Sáng + N sub-row Chiều
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schemas import Assignment, Classroom, ScheduledSession


# --- Hardcoded mapping mã khoa → tên viết tắt hiển thị ---
_DEPT_FULL_NAME = {
    "cntt": "CÔNG NGHỆ THÔNG TIN",
    "kt":   "KINH TẾ",
    "dt":   "ĐIỆN TỬ",
    "co":   "CƠ KHÍ",
    "od":   "Ô TÔ",
    "kg":   "KHOA HỌC CƠ BẢN",
    "tc":   "TÀI CHÍNH",
    "qtkd": "QUẢN TRỊ KINH DOANH",
    "nn":   "NGOẠI NGỮ",
    "may":  "MAY",
    "dtm":  "DU LỊCH - THƯƠNG MẠI",
}

_DAY_VI = {2: "THỨ 2", 3: "THỨ 3", 4: "THỨ 4", 5: "THỨ 5", 6: "THỨ 6", 7: "THỨ 7", 8: "CHỦ NHẬT"}

_REASON_LABELS = {
    "thi":            "Tuần thi",
    "du_phong":       "Dự phòng",
    "quan_su":        "Quân sự",
    "nghi_le":        "Nghỉ lễ",
    "thi_lai":        "Thi lại",
    "thi_tot_nghiep": "Thi tốt nghiệp",
    "thuc_tap":       "Thực tập",
    "tttn":           "Thực tập tốt nghiệp",
    "nghi":           "Nghỉ",
    "thuc_te":        "Thực tế/Thực tập",
    "khac":           "Khác",
}

# --- Calendar status: code + fill + font color ---
# Codes: H=học, N=nghỉ lễ, T=thi, T2=thi lần 2, TL=thi lại, TL2=thi lại lần 2,
#        DP=dự phòng, QS=quân sự, TT=thực tập, TG=tốt nghiệp, K=khác,
#        '—'=chưa bắt đầu, 'Hè'=nghỉ hè
_REASON_TO_CODE = {
    "thi":            "T",
    "thi_lai":        "TL",
    "thi_tot_nghiep": "TG",
    "nghi_le":        "N",
    "nghi":           "N",
    "du_phong":       "DP",
    "quan_su":        "QS",
    "thuc_tap":       "TT",
    "tttn":           "TT",
    "thuc_te":        "TT",
    "khac":           "K",
}

_FILL_H        = PatternFill("solid", fgColor="FFFFFF")     # Học — trắng
_FILL_HOLIDAY  = PatternFill("solid", fgColor="FFF2CC")     # Nghỉ lễ — vàng nhạt
_FILL_EXAM     = PatternFill("solid", fgColor="FF0000")     # Thi — đỏ
_FILL_RETEST   = PatternFill("solid", fgColor="C0C0C0")     # Thi lại — xám
_FILL_GRADUATE = PatternFill("solid", fgColor="FFC000")     # Tốt nghiệp — cam
_FILL_RESERVE  = PatternFill("solid", fgColor="B4C7E7")     # Dự phòng — xanh nhạt
_FILL_MILITARY = PatternFill("solid", fgColor="D6E4F0")     # Quân sự — xanh nước biển
_FILL_INTERN   = PatternFill("solid", fgColor="E2EFDA")     # Thực tập — xanh lá nhạt
_FILL_OTHER    = PatternFill("solid", fgColor="EDEDED")     # Khác — xám nhạt
_FILL_NOT_YET  = PatternFill("solid", fgColor="F2F2F2")     # Chưa bắt đầu — xám rất nhạt
_FILL_SUMMER   = PatternFill("solid", fgColor="FFFFCC")     # Nghỉ hè — vàng nhạt hơn

_CODE_TO_FILL = {
    "H":   _FILL_H,
    "N":   _FILL_HOLIDAY,
    "T":   _FILL_EXAM,
    "TL":  _FILL_RETEST,
    "TG":  _FILL_GRADUATE,
    "DP":  _FILL_RESERVE,
    "QS":  _FILL_MILITARY,
    "TT":  _FILL_INTERN,
    "K":   _FILL_OTHER,
    "—":   _FILL_NOT_YET,
    "Hè":  _FILL_SUMMER,
}

_CODE_TO_FONT_COLOR = {
    "T":  "FFFFFF",   # đỏ → chữ trắng
    "N":  "C00000",   # vàng → chữ đỏ đậm
}

# --- Styles ---
_THIN = Side(style="thin", color="888888")
_MEDIUM = Side(style="medium", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_BOLD = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_BANNER_FILL = PatternFill("solid", fgColor="FFFF00")
_DAY_FILL = PatternFill("solid", fgColor="FCE4D6")
_PERIOD_FILL = PatternFill("solid", fgColor="FFF2CC")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _safe_sheet_name(raw: str, used: set) -> str:
    s = raw or "lop"
    for ch in r":\/?*[]":
        s = s.replace(ch, "_")
    s = s[:31].strip() or "lop"
    base = s
    n = 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = (base[: 31 - len(suffix)]) + suffix
    used.add(s)
    return s


def _format_date(s: str) -> str:
    """YYYY-MM-DD → DD/MM/YYYY."""
    if not s or len(s) < 10:
        return s
    try:
        y, m, d = s[:10].split("-")
        return f"{d}/{m}/{y}"
    except ValueError:
        return s


def _format_week_ranges(weeks: List[int]) -> str:
    """[24,26,27,28,31] → '24, 26-28, 31'."""
    if not weeks:
        return ""
    ranges = []
    start = end = weeks[0]
    for w in weeks[1:]:
        if w == end + 1:
            end = w
        else:
            ranges.append(str(start) if start == end else f"{start}-{end}")
            start = end = w
    ranges.append(str(start) if start == end else f"{start}-{end}")
    return ", ".join(ranges)


# --- Term metadata helpers ---

def _parse_term_code(term_code: str) -> Tuple[str, str]:
    """'2025_2026_HK2' → ('II', '2025 - 2026'). Fallback: ('', term_code)."""
    parts = (term_code or "").split("_")
    if len(parts) >= 3 and parts[2].upper().startswith("HK"):
        hk_num = parts[2][2:].strip()
        roman = {"1": "I", "2": "II", "3": "III"}.get(hk_num, hk_num)
        year = f"{parts[0]} - {parts[1]}"
        return roman, year
    return "", term_code


def _term_week_range(week_dates: Dict[int, Tuple[str, str]]) -> Tuple[int, int, str, str]:
    """Trả về (week_start, week_end, date_start, date_end)."""
    if not week_dates:
        return 0, 0, "", ""
    ws = min(week_dates.keys())
    we = max(week_dates.keys())
    return ws, we, week_dates[ws][0], week_dates[we][1]


def _holiday_note(week_dates: Dict[int, Tuple[str, str]],
                  holiday_reasons: Dict[int, str]) -> str:
    """[Đã bỏ — header compact không hiển thị ghi chú]."""
    return ""


# --- Static header builders (compact) ---

def _draw_static_header(ws, dept_full_name: str, term_code: str,
                        week_dates: Dict[int, Tuple[str, str]],
                        holiday_reasons: Dict[int, str],
                        n_cols: int) -> int:
    """Vẽ header 3 dòng compact. Trả về row tiếp theo (sau header).

    Row 1: trường (trái) | CHXHCN (phải)
    Row 2: KHOA: ... (trái) | THỜI KHÓA BIỂU - HỌC KỲ X NĂM HỌC ... (phải)
    Row 3: Tuần X-Y (dd/mm/yyyy - dd/mm/yyyy)
    """
    hk_roman, year = _parse_term_code(term_code)
    ws_num, we_num, d_start, d_end = _term_week_range(week_dates)

    half = max(1, n_cols // 2)

    # Row 1
    c = ws.cell(row=1, column=1,
                value="BỘ GIÁO DỤC VÀ ĐÀO TẠO\nTRƯỜNG CAO ĐẲNG KỸ THUẬT - CÔNG NGHỆ BÁCH KHOA")
    c.font = Font(bold=True, size=10)
    c.alignment = _CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=half)

    c = ws.cell(row=1, column=half + 1,
                value="CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM\nĐộc lập - Tự do - Hạnh phúc")
    c.font = Font(bold=True, size=10)
    c.alignment = _CENTER
    ws.merge_cells(start_row=1, start_column=half + 1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 36

    # Row 2: KHOA (trái) + tiêu đề (phải)
    c = ws.cell(row=2, column=1, value=f"KHOA: {dept_full_name or ''}")
    c.font = Font(bold=True, size=11)
    c.alignment = _LEFT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=half)

    title = f"THỜI KHÓA BIỂU — HỌC KỲ {hk_roman} NĂM HỌC {year}".strip()
    c = ws.cell(row=2, column=half + 1, value=title)
    c.font = Font(bold=True, size=13)
    c.alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=half + 1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 22

    # Row 3: phạm vi tuần + ngày
    range_text = (
        f"Từ tuần {ws_num} đến tuần {we_num} "
        f"({_format_date(d_start)} → {_format_date(d_end)})"
    ) if ws_num else ""
    c = ws.cell(row=3, column=1, value=range_text)
    c.font = Font(italic=True, size=10)
    c.alignment = _CENTER
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    ws.row_dimensions[3].height = 18

    return 4  # row tiếp theo


# --- Calendar (5 rows: Tháng | Tuần | Từ ngày | Đến ngày | Trạng thái) ---

def _parse_iso(s: str) -> Optional[datetime]:
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d")
    except (ValueError, AttributeError, TypeError):
        return None


def _week_to_month_label(from_date: str) -> str:
    """Tuần thuộc tháng nào → 'M/YYYY'. Dùng giữa tuần (from + 3) để cross-month đẹp."""
    dt = _parse_iso(from_date)
    if not dt:
        return ""
    from datetime import timedelta
    mid = dt + timedelta(days=3)
    return f"{mid.month}/{mid.year}"


def _short_dm(s: str) -> str:
    """YYYY-MM-DD → 'dd/mm' (ngắn gọn cho calendar)."""
    dt = _parse_iso(s)
    if not dt:
        return ""
    return f"{dt.day:02d}/{dt.month:02d}"


def _reason_to_code(raw: str) -> str:
    """Map reason → code calendar.

    Hỗ trợ cả raw text tiếng Việt ('Tết Nguyên Đán', 'Nghỉ lễ 30/4', ...)
    và code chuẩn ('thi', 'nghi_le', ...).
    """
    if not raw:
        return "K"
    # Match canonical code trước
    if raw in _REASON_TO_CODE:
        return _REASON_TO_CODE[raw]
    # Match từ khóa trong text VN
    low = raw.lower()
    if "tết" in low or "tet" in low or "nghỉ lễ" in low or "nghi le" in low or low.startswith("nghỉ"):
        return "N"
    if "thi lại" in low or "thi lai" in low:
        return "TL"
    if "tốt nghiệp" in low or "tot nghiep" in low:
        return "TG"
    if low.startswith("thi") or "thi cuối kỳ" in low or "kết thúc môn" in low:
        return "T"
    if "dự phòng" in low or "du phong" in low:
        return "DP"
    if "quân sự" in low or "quan su" in low:
        return "QS"
    if "thực tập" in low or "thuc tap" in low or "thực tế" in low:
        return "TT"
    return "K"


def _class_week_status(
    week: int,
    holiday_reasons: Dict[int, str],
    class_excl: Dict[int, str],
    class_week_start: Optional[int],
    class_week_end: Optional[int],
) -> Tuple[str, str]:
    """Trả về (code, label) cho 1 tuần.

    Ưu tiên: holiday > excluded riêng lớp > trước class_week_start > Học.

    Lưu ý: KHÔNG suy ra "Hè" từ session.week_end vì sessions chỉ phản ánh tuần
    có môn được xếp, không bao gồm tuần thi/dự phòng/học môn khác. Tuần "Hè"
    chỉ được hiển thị nếu DB cung cấp explicit qua holiday_reasons.
    """
    if week in holiday_reasons:
        r = holiday_reasons[week]
        return _reason_to_code(r), _REASON_LABELS.get(r, r)
    if week in class_excl:
        r = class_excl[week]
        return _reason_to_code(r), _REASON_LABELS.get(r, r)
    if class_week_start is not None and week < class_week_start:
        return "—", "Chưa bắt đầu"
    return "H", "Học"


def _draw_calendar(
    ws,
    start_row: int,
    week_dates: Dict[int, Tuple[str, str]],
    holiday_reasons: Dict[int, str],
    class_excl: Dict[int, str],
    class_week_start: Optional[int],
    class_week_end: Optional[int],
    cal_col_start: int = 1,
) -> int:
    """Vẽ calendar 5 dòng. Mỗi tuần = 1 cột. Trả về row tiếp theo.

    Layout:
      row 0: Tháng (merge các tuần cùng tháng)
      row 1: Số tuần (week_order)
      row 2: Ngày bắt đầu (dd/mm)
      row 3: Ngày kết thúc (dd/mm)
      row 4: Trạng thái (H, N, T, DP, ...)
    """
    if not week_dates:
        return start_row

    weeks_sorted = sorted(week_dates.keys())
    n_weeks = len(weeks_sorted)
    if n_weeks == 0:
        return start_row

    r_month  = start_row
    r_week   = start_row + 1
    r_from   = start_row + 2
    r_to     = start_row + 3
    r_status = start_row + 4

    # Row Tháng — group consecutive weeks cùng tháng và merge
    month_groups: List[Tuple[str, int, int]] = []  # (month_label, col_start, col_end)
    cur_label = None
    cur_start_idx = 0
    for i, w in enumerate(weeks_sorted):
        from_date = week_dates[w][0]
        m = _week_to_month_label(from_date)
        if cur_label is None:
            cur_label = m
            cur_start_idx = 0
        elif m != cur_label:
            month_groups.append((cur_label, cur_start_idx, i - 1))
            cur_label = m
            cur_start_idx = i
    if cur_label is not None:
        month_groups.append((cur_label, cur_start_idx, n_weeks - 1))

    for label, idx_s, idx_e in month_groups:
        col_s = cal_col_start + idx_s
        col_e = cal_col_start + idx_e
        c = ws.cell(row=r_month, column=col_s, value=label)
        c.font = Font(bold=True, size=10)
        c.alignment = _CENTER
        c.border = _BORDER
        c.fill = _HEADER_FILL
        if col_e > col_s:
            ws.merge_cells(start_row=r_month, start_column=col_s,
                           end_row=r_month, end_column=col_e)
        # Border cho tất cả cell trong vùng merge
        for cc in range(col_s, col_e + 1):
            ws.cell(row=r_month, column=cc).border = _BORDER

    # Rows Tuần / Từ / Đến / Trạng thái — 1 cell per week
    for i, w in enumerate(weeks_sorted):
        col = cal_col_start + i
        from_date, to_date = week_dates[w]

        # Số tuần
        c = ws.cell(row=r_week, column=col, value=str(w))
        c.font = Font(size=9)
        c.alignment = _CENTER
        c.border = _BORDER

        # Ngày bắt đầu
        c = ws.cell(row=r_from, column=col, value=_short_dm(from_date))
        c.font = Font(size=9)
        c.alignment = _CENTER
        c.border = _BORDER

        # Ngày kết thúc
        c = ws.cell(row=r_to, column=col, value=_short_dm(to_date))
        c.font = Font(size=9)
        c.alignment = _CENTER
        c.border = _BORDER

        # Trạng thái
        code, _label = _class_week_status(
            w, holiday_reasons, class_excl, class_week_start, class_week_end,
        )
        c = ws.cell(row=r_status, column=col, value=code)
        font_color = _CODE_TO_FONT_COLOR.get(code, "000000")
        c.font = Font(size=9, bold=True, color=font_color)
        c.alignment = _CENTER
        c.border = _BORDER
        c.fill = _CODE_TO_FILL.get(code, _FILL_H)

    # Set heights
    for r in (r_month, r_week, r_from, r_to, r_status):
        ws.row_dimensions[r].height = 16

    return start_row + 5  # row tiếp theo


# --- Exam banner (row 14) ---

def _build_exam_banner(class_excl: Dict[int, str]) -> str:
    """Sinh banner thi từ excluded_weeks của lớp.

    Gom 3 reason: thi (KTHM), thi_lai, thi_tot_nghiep.
    """
    by_reason: Dict[str, List[int]] = {"thi": [], "thi_lai": [], "thi_tot_nghiep": []}
    for w, r in class_excl.items():
        if r in by_reason:
            by_reason[r].append(int(w))

    parts = []
    if by_reason["thi"]:
        parts.append(
            f"LỊCH THI KẾT THÚC MÔN HỌC/MÔ ĐUN DỰ KIẾN VÀO TUẦN {_format_week_ranges(sorted(by_reason['thi']))}"
        )
    if by_reason["thi_lai"]:
        parts.append(
            f"LỊCH THI LẠI DỰ KIẾN VÀO TUẦN {_format_week_ranges(sorted(by_reason['thi_lai']))}"
        )
    if by_reason["thi_tot_nghiep"]:
        parts.append(
            f"LỊCH TỐT NGHIỆP TUẦN {_format_week_ranges(sorted(by_reason['thi_tot_nghiep']))}"
        )
    return ", ".join(parts)


def _draw_exam_banner(ws, row: int, text: str, n_cols: int) -> None:
    c = ws.cell(row=row, column=1, value=text or "")
    c.font = Font(bold=True, size=11, color="C00000")
    c.alignment = _CENTER
    c.fill = _BANNER_FILL
    c.border = _BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 36


# --- Grid header (rows 15-17) ---

def _draw_grid_header(ws, header_row: int, days: List[int],
                      tiet_per_day: int = 5) -> None:
    """Vẽ 3 dòng header: TÊN LỚP | BUỔI | THỨ × tiet_per_day cột.

    Cột bố cục:
      col 1 = TÊN LỚP (merge 3 rows)
      col 2 = BUỔI    (merge 3 rows)
      col 3..        = THỨ 2..CN, mỗi thứ tiet_per_day cột
    """
    r1, r2, r3 = header_row, header_row + 1, header_row + 2

    # TÊN LỚP
    c = ws.cell(row=r1, column=1, value="TÊN LỚP")
    c.font = Font(bold=True, size=10)
    c.alignment = _CENTER
    c.fill = _HEADER_FILL
    c.border = _BORDER
    ws.merge_cells(start_row=r1, start_column=1, end_row=r3, end_column=1)

    # BUỔI
    c = ws.cell(row=r1, column=2, value="BUỔI")
    c.font = Font(bold=True, size=10)
    c.alignment = _CENTER
    c.fill = _HEADER_FILL
    c.border = _BORDER
    ws.merge_cells(start_row=r1, start_column=2, end_row=r3, end_column=2)

    # THỨ
    for i, d in enumerate(days):
        col_start = 3 + i * tiet_per_day
        col_end = col_start + tiet_per_day - 1

        c = ws.cell(row=r1, column=col_start, value=_DAY_VI.get(int(d), str(d)))
        c.font = Font(bold=True, size=10)
        c.alignment = _CENTER
        c.fill = _DAY_FILL
        c.border = _BORDER
        ws.merge_cells(start_row=r1, start_column=col_start, end_row=r1, end_column=col_end)

        c = ws.cell(row=r2, column=col_start, value="Môn học/ Giáo viên/ Tuần/Phòng học")
        c.font = Font(bold=True, italic=True, size=9)
        c.alignment = _CENTER
        c.fill = _HEADER_FILL
        c.border = _BORDER
        ws.merge_cells(start_row=r2, start_column=col_start, end_row=r2, end_column=col_end)

        for k in range(tiet_per_day):
            cc = ws.cell(row=r3, column=col_start + k, value=f"Tiết {k + 1}")
            cc.font = Font(bold=True, size=9)
            cc.alignment = _CENTER
            cc.fill = _PERIOD_FILL
            cc.border = _BORDER

    ws.row_dimensions[r1].height = 22
    ws.row_dimensions[r2].height = 26
    ws.row_dimensions[r3].height = 18


# --- Cell content builder (5-line format) ---

def _build_cell_text_pdt(
    s: ScheduledSession,
    a: Optional[Assignment],
    label: Dict[str, str],
    classrooms_by_id: Dict[int, Classroom],
    week_dates: Dict[int, Tuple[str, str]],
    holiday_reasons: Dict[int, str],
    class_excluded_weeks: Dict[str, Dict[int, str]],
) -> str:
    """Format cell theo mẫu phòng đào tạo:

      MĐ32: Phát triển Web ... (80 tiết)
      GV: Nguyễn Thị Luyện
      Tuần 32-42 (09/03/2026 → 24/05/2026)
      Nghỉ tuần 35, 36
      P. 204 - N1
    """
    sub_code = label.get("subject_code", "")
    sub_name = label.get("subject_name", "")
    total_hours = label.get("total_hours", "")
    teacher_name = label.get("teacher_name", s.teacher_id)

    room = classrooms_by_id.get(int(s.classroom_id))
    room_label = (room.name if room else None) or str(s.classroom_id)

    w_start = int(s.week_start)
    w_end = int(s.week_end)

    # Lấy teaching/skipped từ session nếu có (solver đã set), else fallback
    if s.teaching_weeks:
        teaching_weeks = list(s.teaching_weeks)
        all_skipped: Dict[int, str] = dict(s.skipped_weeks)
    else:
        cls_id = label.get("class_id", "")
        class_excl = class_excluded_weeks.get(cls_id, {})
        excl_in_range = {w: r for w, r in class_excl.items() if w_start <= w <= w_end}
        holidays_in_range = {
            w: holiday_reasons.get(w, "Nghỉ")
            for w in range(w_start, w_end + 1)
            if w in holiday_reasons
        }
        all_skipped = {**holidays_in_range, **excl_in_range}
        teaching_weeks = sorted(w for w in range(w_start, w_end + 1) if w not in all_skipped)

    teaching_range = _format_week_ranges(teaching_weeks)

    # Dòng tuần — format mẫu phòng đào tạo: Tuần 24-32 (12/01/2026 → 15/03/2026)
    if teaching_weeks:
        fd = _format_date(week_dates.get(teaching_weeks[0], ("", ""))[0])
        td = _format_date(week_dates.get(teaching_weeks[-1], ("", ""))[1])
        week_line = f"Tuần {teaching_range} ({fd} → {td})" if fd and td else f"Tuần {teaching_range}"
    else:
        week_line = f"Tuần {w_start}-{w_end}"

    # Dòng nghỉ — gom các tuần skipped (chỉ hiển thị nếu có trong khoảng teaching)
    skipped_text = ""
    if all_skipped:
        skipped_weeks_sorted = sorted(all_skipped.keys())
        skipped_text = f"Nghỉ tuần {_format_week_ranges(skipped_weeks_sorted)}"

    # Tên môn: "MĐ32: Phát triển Web ... (XX tiết)"
    name_part = f"{sub_code}: {sub_name}" if sub_code else sub_name
    if total_hours:
        name_part += f" ({total_hours} tiết)"

    parts = [
        name_part,
        f"GV: {teacher_name}",
        week_line,
    ]
    if skipped_text:
        parts.append(skipped_text)
    parts.append(f"P. {room_label}")

    return "\n".join(parts)


# --- Layout sessions per class block ---

def _layout_class_sessions(
    cls_sessions: List[ScheduledSession],
    days: List[int],
    last_morning: int,
) -> Tuple[Dict[int, List[ScheduledSession]],
           Dict[int, List[ScheduledSession]],
           int, int]:
    """Tách sessions theo buổi (Sáng/Chiều) và theo thứ.

    Returns: (sang_by_day, chieu_by_day, n_sub_sang, n_sub_chieu).
    """
    sang: Dict[int, List[ScheduledSession]] = {d: [] for d in days}
    chieu: Dict[int, List[ScheduledSession]] = {d: [] for d in days}

    for s in cls_sessions:
        d = int(s.day)
        if d not in sang:
            sang[d] = []
            chieu[d] = []
        if int(s.period_start) <= last_morning:
            sang[d].append(s)
        else:
            chieu[d].append(s)

    n_sang = max((len(v) for v in sang.values()), default=0) or 1
    n_chieu = max((len(v) for v in chieu.values()), default=0) or 1
    return sang, chieu, n_sang, n_chieu


def _write_session_cell(ws, row: int, col_start: int, col_end: int,
                        row_span: int, text: str) -> None:
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(size=9)
    c.alignment = _CENTER
    c.border = _BORDER
    if col_end > col_start or row_span > 1:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row + row_span - 1, end_column=col_end,
        )
    # Border cho mọi cell trong vùng merge
    for r in range(row, row + row_span):
        for cc in range(col_start, col_end + 1):
            ws.cell(row=r, column=cc).border = _BORDER


def _write_class_block(
    ws,
    start_row: int,
    cls_name: str,
    cls_sessions: List[ScheduledSession],
    a_by_id: Dict[int, Assignment],
    assignment_labels: Dict[int, Dict[str, str]],
    classrooms_by_id: Dict[int, Classroom],
    days: List[int],
    last_morning: int,
    week_dates: Dict[int, Tuple[str, str]],
    holiday_reasons: Dict[int, str],
    class_excluded_weeks: Dict[str, Dict[int, str]],
    tiet_per_day: int = 5,
) -> int:
    """Vẽ 1 block lớp tại start_row. Trả về row tiếp theo (sau block)."""
    sang, chieu, n_sang, n_chieu = _layout_class_sessions(cls_sessions, days, last_morning)
    total_rows = n_sang + n_chieu

    # Cột TÊN LỚP (col 1) — merge dọc toàn block
    c = ws.cell(row=start_row, column=1, value=cls_name)
    c.font = Font(bold=True, size=11)
    c.alignment = _CENTER
    c.border = _BORDER
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row + total_rows - 1, end_column=1,
    )

    # Cột BUỔI (col 2)
    c = ws.cell(row=start_row, column=2, value="Sáng")
    c.font = Font(bold=True, size=10)
    c.alignment = _CENTER
    c.border = _BORDER
    if n_sang > 1:
        ws.merge_cells(
            start_row=start_row, start_column=2,
            end_row=start_row + n_sang - 1, end_column=2,
        )

    chieu_row = start_row + n_sang
    c = ws.cell(row=chieu_row, column=2, value="Chiều")
    c.font = Font(bold=True, size=10)
    c.alignment = _CENTER
    c.border = _BORDER
    if n_chieu > 1:
        ws.merge_cells(
            start_row=chieu_row, start_column=2,
            end_row=chieu_row + n_chieu - 1, end_column=2,
        )

    # Vẽ lưới session
    def _fill_session_grid(buoi_start_row: int, buoi_sessions: Dict[int, List[ScheduledSession]],
                           n_sub: int) -> None:
        for i, d in enumerate(days):
            col_start = 3 + i * tiet_per_day
            col_end = col_start + tiet_per_day - 1
            sess_list = buoi_sessions.get(d, [])
            if not sess_list:
                # Empty cell — vẽ border
                for r in range(buoi_start_row, buoi_start_row + n_sub):
                    for cc in range(col_start, col_end + 1):
                        ws.cell(row=r, column=cc).border = _BORDER
                if n_sub > 1:
                    ws.merge_cells(
                        start_row=buoi_start_row, start_column=col_start,
                        end_row=buoi_start_row + n_sub - 1, end_column=col_end,
                    )
                continue

            if len(sess_list) == 1:
                # 1 môn → merge full vùng
                s = sess_list[0]
                a = a_by_id.get(int(s.assignment_id))
                lab = assignment_labels.get(int(s.assignment_id), {})
                text = _build_cell_text_pdt(
                    s, a, lab, classrooms_by_id,
                    week_dates, holiday_reasons, class_excluded_weeks,
                )
                _write_session_cell(ws, buoi_start_row, col_start, col_end, n_sub, text)
            else:
                # k ≥ 2 môn → mỗi môn 1 sub-row, sub-row dư ở cuối ghép với môn cuối
                k = len(sess_list)
                # Phân bổ: k-1 môn đầu mỗi môn 1 row, môn cuối lấy phần dư
                for j, s in enumerate(sess_list):
                    a = a_by_id.get(int(s.assignment_id))
                    lab = assignment_labels.get(int(s.assignment_id), {})
                    text = _build_cell_text_pdt(
                        s, a, lab, classrooms_by_id,
                        week_dates, holiday_reasons, class_excluded_weeks,
                    )
                    sub_row = buoi_start_row + j
                    if j == k - 1 and n_sub > k:
                        span = n_sub - j
                    else:
                        span = 1
                    _write_session_cell(ws, sub_row, col_start, col_end, span, text)

    _fill_session_grid(start_row, sang, n_sang)
    _fill_session_grid(chieu_row, chieu, n_chieu)

    # Set chiều cao row data
    for r in range(start_row, start_row + total_rows):
        ws.row_dimensions[r].height = 80

    return start_row + total_rows


def export_timetable_by_class_pdt(
    sessions: List[ScheduledSession],
    assignments: List[Assignment],
    assignment_labels: Dict[int, Dict[str, str]],
    classrooms: List[Classroom],
    *,
    days: List[int],
    periods_per_day: int,
    morning_periods: Optional[List[int]],
    xlsx_path: Path,
    term_code: str = "",
    department_code: str = "",
    week_dates: Optional[Dict[int, Tuple[str, str]]] = None,
    holiday_reasons: Optional[Dict[int, str]] = None,
    class_excluded_weeks: Optional[Dict[str, Dict[int, str]]] = None,
    class_week_starts: Optional[Dict[str, int]] = None,
) -> None:
    """Xuất Excel theo mẫu phòng đào tạo: 1 file, mỗi lớp 1 sheet."""
    a_by_id = {int(a.id): a for a in assignments}
    room_by_id = {int(c.id): c for c in classrooms}
    week_dates = week_dates or {}
    holiday_reasons = holiday_reasons or {}
    class_excluded_weeks = class_excluded_weeks or {}
    class_week_starts = class_week_starts or {}

    last_morning = max(morning_periods) if morning_periods else periods_per_day // 2
    dept_full_name = _DEPT_FULL_NAME.get((department_code or "").lower(), (department_code or "").upper())

    # Group sessions theo lớp
    by_class: Dict[str, List[ScheduledSession]] = {}
    for s in sessions:
        lab = assignment_labels.get(int(s.assignment_id), {})
        cls = str(lab.get("class_id", "")).strip() or f"unknown_{s.assignment_id}"
        by_class.setdefault(cls, []).append(s)

    if not by_class:
        return

    tiet_per_day = 5
    n_cols = 2 + len(days) * tiet_per_day

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()

    for cls_name in sorted(by_class.keys()):
        cls_sessions = by_class[cls_name]
        sheet_name = _safe_sheet_name(cls_name, used_names)
        ws = wb.create_sheet(title=sheet_name)

        # Static header (rows 1-3) → trả về row tiếp theo
        next_row = _draw_static_header(
            ws, dept_full_name, term_code,
            week_dates, holiday_reasons, n_cols,
        )

        # Calendar (rows 4-8): 5 rows × n_cols (1 cột/tuần, fit từ col 1 đến n_cols)
        cls_excl = class_excluded_weeks.get(cls_name, {})
        cls_week_start = class_week_starts.get(cls_name)
        cls_week_end = max((int(s.week_end) for s in cls_sessions), default=None)

        # Calendar dùng N cột đầu tiên = số tuần. Nếu n_weeks > n_cols, vẫn vẽ — Excel auto co giãn
        cal_end_row = _draw_calendar(
            ws,
            start_row=next_row,
            week_dates=week_dates,
            holiday_reasons=holiday_reasons,
            class_excl=cls_excl,
            class_week_start=cls_week_start,
            class_week_end=cls_week_end,
            cal_col_start=1,
        )

        # Spacer 1 row
        spacer_row = cal_end_row
        ws.row_dimensions[spacer_row].height = 6

        # Exam banner
        banner_row = spacer_row + 1
        banner_text = _build_exam_banner(cls_excl)
        _draw_exam_banner(ws, banner_row, banner_text, n_cols)

        # Grid header (3 rows)
        grid_header_row = banner_row + 1
        _draw_grid_header(ws, header_row=grid_header_row, days=days, tiet_per_day=tiet_per_day)

        # Class block
        data_start_row = grid_header_row + 3
        _write_class_block(
            ws,
            start_row=data_start_row,
            cls_name=cls_name,
            cls_sessions=cls_sessions,
            a_by_id=a_by_id,
            assignment_labels=assignment_labels,
            classrooms_by_id=room_by_id,
            days=days,
            last_morning=last_morning,
            week_dates=week_dates,
            holiday_reasons=holiday_reasons,
            class_excluded_weeks=class_excluded_weeks,
            tiet_per_day=tiet_per_day,
        )

        # Column widths
        ws.column_dimensions["A"].width = 8   # TÊN LỚP / Calendar col 1
        ws.column_dimensions["B"].width = 6   # BUỔI / Calendar col 2
        for i in range(len(days)):
            for k in range(tiet_per_day):
                col_letter = get_column_letter(3 + i * tiet_per_day + k)
                ws.column_dimensions[col_letter].width = 7

        ws.freeze_panes = ws.cell(row=data_start_row, column=3)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
