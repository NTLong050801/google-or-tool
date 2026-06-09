"""Export TKB theo giảng viên: 1 file Excel, mỗi GV 1 sheet (lưới Tiết × Thứ)."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schemas import Assignment, Classroom, ScheduledSession

_DAY_VI = {2: "Thứ 2", 3: "Thứ 3", 4: "Thứ 4", 5: "Thứ 5", 6: "Thứ 6", 7: "Thứ 7", 8: "Chủ nhật"}

_PROGRAM_LABEL = {
    "trung_cap": "Trung cấp",
    "cao_dang":  "Cao đẳng",
    "lien_thong": "Liên thông",
}

_REASON_LABELS: dict = {
    "thi":      "Tuần thi",
    "du_phong": "Dự phòng",
    "quan_su":  "Quân sự",
    "nghi_le":  "Nghỉ lễ",
    "thi_lai":  "Thi lại",
    "nghi":     "Nghỉ",
    "thuc_te":  "Thực tế/Thực tập",
    "khac":     "Khác",
}

_THIN          = Side(style="thin", color="888888")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL   = PatternFill("solid", fgColor="D9E1F2")
_MORNING_FILL  = PatternFill("solid", fgColor="FFF2CC")
_AFTERNOON_FILL= PatternFill("solid", fgColor="E2EFDA")
_CONFLICT_FILL = PatternFill("solid", fgColor="F8CBAD")
_CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _safe_sheet_name(raw: str, used: set) -> str:
    s = raw or "gv"
    for ch in r":\/?*[]":
        s = s.replace(ch, "_")
    s = s[:31].strip() or "gv"
    base, n = s, 1
    while s in used:
        n += 1
        suffix = f"_{n}"
        s = base[: 31 - len(suffix)] + suffix
    used.add(s)
    return s


def _format_date(s: str) -> str:
    if not s or len(s) < 10:
        return s
    try:
        y, m, d = s[:10].split("-")
        return f"{d}/{m}/{y}"
    except ValueError:
        return s


def _format_week_ranges(weeks: list) -> str:
    if not weeks:
        return ""
    ranges, start, end = [], weeks[0], weeks[0]
    for w in weeks[1:]:
        if w == end + 1:
            end = w
        else:
            ranges.append(str(start) if start == end else f"{start}-{end}")
            start = end = w
    ranges.append(str(start) if start == end else f"{start}-{end}")
    return ", ".join(ranges)


def _build_cell_text(
    s: ScheduledSession,
    a: Optional[Assignment],
    label: Dict[str, str],
    classrooms_by_id: Dict[int, Classroom],
    week_dates: Dict[int, Tuple[str, str]],
    holiday_reasons: Dict[int, str],
    class_excluded_weeks: Dict[str, Dict[int, str]],
) -> str:
    sub_code  = label.get("subject_code", "")
    sub_name  = label.get("subject_name", "")
    cls_id    = label.get("class_id", "")
    program   = label.get("program_level", "")
    cluster   = int(a.lessons_cluster) if a else (int(s.period_end) - int(s.period_start) + 1)

    room = classrooms_by_id.get(int(s.classroom_id))
    room_label = (room.name if room else None) or str(s.classroom_id)

    w_start = int(s.week_start)
    w_end   = int(s.week_end)

    if s.teaching_weeks:
        teaching_weeks = s.teaching_weeks
        all_skipped: Dict[int, str] = dict(s.skipped_weeks)
    else:
        class_excl     = class_excluded_weeks.get(cls_id, {})
        excl_in_range  = {w: r for w, r in class_excl.items() if w_start <= w <= w_end}
        holidays_range = {
            w: holiday_reasons.get(w, "Nghỉ")
            for w in range(w_start, w_end + 1)
            if w in holiday_reasons
        }
        all_skipped    = {**holidays_range, **excl_in_range}
        teaching_weeks = sorted(w for w in range(w_start, w_end + 1) if w not in all_skipped)

    teaching_range = _format_week_ranges(teaching_weeks)
    if teaching_weeks:
        fd, _ = week_dates.get(teaching_weeks[0],  ("", ""))
        _, td = week_dates.get(teaching_weeks[-1], ("", ""))
        week_line = f"Tuần {teaching_range}"
        if fd and td:
            week_line += f" ({_format_date(fd)} → {_format_date(td)})"
        week_line += f"  [{len(teaching_weeks)} tuần]"
    else:
        week_line = f"Tuần {w_start}-{w_end}  [0 tuần]"

    skipped_text = ""
    if all_skipped:
        items = [f"W{w} ({_REASON_LABELS.get(r, r)})" for w, r in sorted(all_skipped.items())]
        skipped_text = "Không dạy: " + " | ".join(items)

    prog_label = _PROGRAM_LABEL.get(program, program or "")
    class_line = cls_id
    if prog_label:
        class_line += f" ({prog_label})"

    parts = [
        f"{sub_code} – {sub_name}" if sub_code else sub_name,
        f"Lớp: {class_line}",
        f"Phòng: {room_label}",
        week_line,
    ]
    if skipped_text:
        parts.append(skipped_text)
    return "\n".join(p for p in parts if p)


def export_timetable_by_teacher(
    sessions: List[ScheduledSession],
    assignments: List[Assignment],
    assignment_labels: Dict[int, Dict[str, str]],
    classrooms: List[Classroom],
    *,
    days: List[int],
    periods_per_day: int,
    morning_periods: Optional[List[int]],
    xlsx_path: Path,
    week_dates: Optional[Dict[int, Tuple[str, str]]] = None,
    holiday_reasons: Optional[Dict[int, str]] = None,
    class_excluded_weeks: Optional[Dict[str, Dict[int, str]]] = None,
) -> int:
    """Xuất Excel TKB theo GV: mỗi sheet = 1 GV, lưới Tiết × Thứ.

    Trả về số GV đã xuất.
    """
    a_by_id    = {int(a.id): a for a in assignments}
    room_by_id = {int(c.id): c for c in classrooms}
    week_dates          = week_dates or {}
    holiday_reasons     = holiday_reasons or {}
    class_excluded_weeks = class_excluded_weeks or {}

    # Gom sessions theo teacher_id
    by_teacher: Dict[str, List[ScheduledSession]] = {}
    for s in sessions:
        tid = str(s.teacher_id or "").strip() or "unknown"
        by_teacher.setdefault(tid, []).append(s)

    if not by_teacher:
        return 0

    last_morning = max(morning_periods) if morning_periods else periods_per_day // 2

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set = set()

    for tid in sorted(by_teacher.keys()):
        t_sessions = by_teacher[tid]

        # Lấy tên GV từ label của session đầu tiên
        first_label = assignment_labels.get(int(t_sessions[0].assignment_id), {})
        teacher_name = first_label.get("teacher_name", tid)
        sheet_title  = teacher_name if teacher_name and teacher_name != tid else tid
        sheet_name   = _safe_sheet_name(sheet_title, used_names)
        ws = wb.create_sheet(title=sheet_name)

        n_cols = 1 + len(days)
        title = f"Thời khóa biểu giảng viên: {teacher_name}"
        if teacher_name != tid:
            title += f" ({tid})"
        ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = _LEFT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

        header_row = 3
        ws.cell(row=header_row, column=1, value="Tiết").font = Font(bold=True)
        ws.cell(row=header_row, column=1).fill = _HEADER_FILL
        ws.cell(row=header_row, column=1).alignment = _CENTER
        ws.cell(row=header_row, column=1).border = _BORDER
        for i, d in enumerate(days):
            c = ws.cell(row=header_row, column=2 + i, value=_DAY_VI.get(int(d), str(d)))
            c.font = Font(bold=True)
            c.fill = _HEADER_FILL
            c.alignment = _CENTER
            c.border = _BORDER

        for p in range(1, periods_per_day + 1):
            r    = header_row + p
            fill = _MORNING_FILL if p <= last_morning else _AFTERNOON_FILL
            cell = ws.cell(row=r, column=1, value=f"Tiết {p}")
            cell.alignment = _CENTER
            cell.border    = _BORDER
            cell.fill      = fill
            for i in range(len(days)):
                c = ws.cell(row=r, column=2 + i)
                c.border    = _BORDER
                c.alignment = _CENTER
                c.fill      = fill

        # Đặt nội dung từng session
        # slot_content: key → list of (content, teaching_weeks_set)
        slot_content: Dict[Tuple[int, int], List[tuple]] = {}
        slot_span:    Dict[Tuple[int, int], int]          = {}
        for s in t_sessions:
            a       = a_by_id.get(int(s.assignment_id))
            cluster = int(a.lessons_cluster) if a else (int(s.period_end) - int(s.period_start) + 1)
            lab     = assignment_labels.get(int(s.assignment_id), {})
            content = _build_cell_text(s, a, lab, room_by_id, week_dates, holiday_reasons, class_excluded_weeks)
            # Tính teaching weeks để phát hiện conflict thực
            if s.teaching_weeks:
                tw = set(s.teaching_weeks)
            else:
                cls_id     = lab.get("class_id", "")
                cls_excl   = class_excluded_weeks.get(cls_id, {})
                w_start, w_end = int(s.week_start), int(s.week_end)
                tw = set(
                    w for w in range(w_start, w_end + 1)
                    if w not in holiday_reasons and w not in cls_excl
                )
            key = (int(s.day), int(s.period_start))
            slot_content.setdefault(key, []).append((content, tw))
            slot_span[key] = cluster

        for (day, p_start), entries in slot_content.items():
            if day not in days:
                continue
            col     = 2 + days.index(day)
            row     = header_row + p_start
            cluster = slot_span[(day, p_start)]
            text    = "\n──────\n".join(e[0] for e in entries)
            cell    = ws.cell(row=row, column=col, value=text)
            cell.alignment = _CENTER
            cell.border    = _BORDER
            # Chỉ tô conflict khi các tuần thực dạy overlap nhau
            if len(entries) > 1:
                all_tw = [e[1] for e in entries]
                has_conflict = any(
                    all_tw[i] & all_tw[j]
                    for i in range(len(all_tw))
                    for j in range(i + 1, len(all_tw))
                )
                if has_conflict:
                    cell.fill = _CONFLICT_FILL
            if cluster > 1:
                ws.merge_cells(
                    start_row=row, start_column=col,
                    end_row=row + cluster - 1, end_column=col,
                )
                for r2 in range(row, row + cluster):
                    ws.cell(row=r2, column=col).border = _BORDER

        ws.column_dimensions["A"].width = 8
        for i in range(len(days)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 36
        for p in range(1, periods_per_day + 1):
            ws.row_dimensions[header_row + p].height = 30
        ws.row_dimensions[header_row].height = 18
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return len(by_teacher)
